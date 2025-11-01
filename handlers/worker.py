# handlers/worker.py

from aiogram import Router, F, types
from aiogram.types import (
    Message, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from datetime import datetime, date, time, timedelta
import os
import pytz
from typing import List

from config import SUPERADMIN_ID
import database
from keyboards.worker_kb import get_worker_kb, get_bonus_kb

# Excel uchun
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = Router()

# ====== Sozlamalar ======
UZ_TZ = pytz.timezone("Asia/Tashkent")
WORK_START = time(9, 0, 0)     # Me'yoriy ish boshlanish vaqti: 09:00
CUTOFF = time(0, 10, 0)        # Kun almashish chegarasi: 00:10
RATE_PER_HOUR = 10_000         # Bonus/Jarima stavkasi: 1 soat = 10 000 so‘m


def _parse_superadmins() -> List[int]:
    """SUPERADMIN_ID: int yoki '1,2,3' bo‘lishi mumkin."""
    try:
        if isinstance(SUPERADMIN_ID, int):
            return [int(SUPERADMIN_ID)]
        txt = str(SUPERADMIN_ID)
        return [int(x.strip()) for x in txt.split(",") if x.strip().isdigit()]
    except Exception:
        return []
SUPERADMINS = _parse_superadmins()


def business_now():
    """Toshkent vaqtida hozirgi vaqt (aware)."""
    return datetime.now(UZ_TZ)


def business_date(now_dt: datetime | None = None) -> date:
    """
    Kun almashish qoidasi:
    - 00:00 .. 00:09:59 — hali kechagi kun hisoblanadi
    - 00:10 .. keyin — yangi kun
    """
    now_dt = now_dt or business_now()
    bdate = now_dt.date()
    if now_dt.time() < CUTOFF:
        return bdate - timedelta(days=1)
    return bdate


def fmt_sum(v) -> str:
    try:
        return f"{float(v):,.0f}".replace(",", " ")
    except:
        return str(v)


def ensure_bonus_tables():
    """bonuses/fines jadvallarini yo‘q bo‘lsa yaratib qo‘yadi."""
    try:
        database.execute("""
            CREATE TABLE IF NOT EXISTS bonuses (
                id SERIAL PRIMARY KEY,
                user_id BIGINT NOT NULL,
                amount NUMERIC NOT NULL,
                reason TEXT,
                created_by BIGINT,
                auto BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        database.execute("""
            CREATE TABLE IF NOT EXISTS fines (
                id SERIAL PRIMARY KEY,
                user_id BIGINT NOT NULL,
                amount NUMERIC NOT NULL,
                reason TEXT,
                created_by BIGINT,
                auto BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
    except Exception:
        # agar hosting CREATE TABLE ni cheklasa ham bot ishlashda davom etsin
        pass


# ============= FSM holatlar =============
class ReportState(StatesGroup):
    income = State()         # 💰 Daromad
    expense = State()        # 💸 Rashod
    product_loop = State()   # 🏪 Mahsulotlarni ketma-ket so‘rash
    confirm = State()        # ✅ Yakun va tasdiqlash


# ===============================
# 🕘 Ishni boshladim (Bonus/Jarima avtomatik)
# ===============================
@router.message(F.text == "🕘 Ishni boshladim")
async def start_work(message: Message):
    ensure_bonus_tables()
    user_id = message.from_user.id
    now = business_now()
    bdate = business_date(now)
    start_time = now.strftime("%H:%M:%S")

    # Bugungi (ish kuni) uchun yozuv bormi?
    existing = database.fetchone(
        "SELECT start_time FROM reports WHERE user_id=:u AND date=:d",
        {"u": user_id, "d": bdate}
    )
    if existing:
        prev_start = existing["start_time"]
        # Oldingi bonus yoki jarimani topamiz
        bonus = database.fetchone(
            "SELECT amount, reason FROM bonuses WHERE user_id=:u AND DATE(created_at)=:d AND auto=TRUE",
            {"u": user_id, "d": bdate}
        )
        fine = database.fetchone(
            "SELECT amount, reason FROM fines WHERE user_id=:u AND DATE(created_at)=:d AND auto=TRUE",
            {"u": user_id, "d": bdate}
        )

        text = f"⚠️ Siz bugungi ishni allaqachon boshlagansiz.\n🕘 Ishni boshlagan vaqt: <b>{prev_start}</b>\n\n"
        if bonus:
            text += f"✅ Bonus: {fmt_sum(bonus['amount'])} so‘m\n🟢 {bonus['reason']}"
        elif fine:
            text += f"❌ Jarima: {fmt_sum(fine['amount'])} so‘m\n🔴 {fine['reason']}"
        else:
            text += "ℹ️ Bonus yoki jarima aniqlanmagan."

        await message.answer(text, parse_mode="HTML")
        return

    # Foydalanuvchi filialini olamiz
    user = database.fetchone("SELECT branch_id, full_name FROM users WHERE telegram_id=:tid", {"tid": user_id})
    branch_id = user["branch_id"] if user else None

    # Ishni boshlashni yozamiz
    database.execute("""
        INSERT INTO reports (user_id, branch_id, date, start_time)
        VALUES (:u, :b, :d, :t)
    """, {"u": user_id, "b": branch_id, "d": bdate, "t": start_time})

    # === Bonus/Jarima hisoblash (09:00 ga nisbatan) ===
    target_dt = datetime.combine(bdate, WORK_START, tzinfo=UZ_TZ)
    diff_minutes = (now - target_dt).total_seconds() / 60
    result_text = f"🕘 Ishni boshladingiz: <b>{start_time}</b>\n"

    if diff_minutes > 10:  # 10 daqiqadan ortiq kechiksa - jarima
        penalty = round((diff_minutes / 60) * RATE_PER_HOUR)
        database.execute("""
            INSERT INTO fines (user_id, amount, reason, created_by, auto)
            VALUES (:u, :a, :r, :c, TRUE)
        """, {
            "u": user_id,
            "a": penalty,
            "r": f"{diff_minutes:.0f} daqiqa kech kelgani uchun avtomatik jarima",
            "c": user_id
        })
        result_text += f"\n❌ Jarima: {fmt_sum(penalty)} so‘m\n🔴 {diff_minutes:.0f} daqiqa kech keldingiz."
    elif diff_minutes < -5:  # 5 daqiqadan ko‘p erta kelsa - bonus
        bonus = round((abs(diff_minutes) / 60) * RATE_PER_HOUR)
        database.execute("""
            INSERT INTO bonuses (user_id, amount, reason, created_by, auto)
            VALUES (:u, :a, :r, :c, TRUE)
        """, {
            "u": user_id,
            "a": bonus,
            "r": f"{abs(diff_minutes):.0f} daqiqa erta kelgani uchun avtomatik bonus",
            "c": user_id
        })
        result_text += f"\n✅ Bonus: {fmt_sum(bonus)} so‘m\n🟢 {abs(diff_minutes):.0f} daqiqa erta keldingiz."
    else:
        result_text += "\n👍 O‘z vaqtida ishga keldingiz."

    await message.answer(result_text, parse_mode="HTML")


# ===============================
# 🏁 Ishni tugatdim (ish kunini yopish)
# ===============================
@router.message(F.text == "🏁 Ishni tugatdim")
async def finish_work(message: Message):
    user_id = message.from_user.id
    now = business_now()
    bdate = business_date(now)
    end_time = now.strftime("%H:%M:%S")

    database.execute(
        "UPDATE reports SET end_time=:t WHERE user_id=:u AND date=:d",
        {"t": end_time, "u": user_id, "d": bdate}
    )

    await message.answer(
        f"🏁 Ish tugash vaqti saqlandi: <b>{end_time}</b>\n\n"
        f"ℹ️ 00:10 dan keyin yangi ish kuni ochiladi.",
        parse_mode="HTML"
    )


# ===============================
# 🧾 Bugungi hisobotni yuborish (FSM)
# ===============================
@router.message(F.text == "🧾 Bugungi hisobotni yuborish")
async def start_report(message: types.Message, state: FSMContext):
    await message.answer("💰 Bugungi daromadni kiriting (faqat raqam):")
    await state.set_state(ReportState.income)


@router.message(ReportState.income)
async def get_income(message: types.Message, state: FSMContext):
    try:
        val = float(message.text.replace(" ", ""))
    except ValueError:
        await message.answer("❌ Faqat raqam kiriting. Masalan: 1200000")
        return

    await state.update_data(income=val)
    await message.answer("💸 Bugungi rashodni kiriting (faqat raqam):")
    await state.set_state(ReportState.expense)


@router.message(ReportState.expense)
async def get_expense(message: types.Message, state: FSMContext):
    try:
        val = float(message.text.replace(" ", ""))
    except ValueError:
        await message.answer("❌ Faqat raqam kiriting. Masalan: 300000")
        return

    await state.update_data(expense=val)

    # Ishchining filialini aniqlaymiz
    user_id = message.from_user.id
    user = database.fetchone("SELECT branch_id FROM users WHERE telegram_id=:tid", {"tid": user_id})
    if not user:
        await message.answer("⚠️ Sizning filial topilmadi.")
        await state.clear()
        return

    branch_id = user["branch_id"]
    products = database.list_products_by_branch(branch_id)

    if not products:
        await message.answer("📦 Omborda mahsulotlar yo‘q — hisobot faqat daromad/rashod bilan saqlanadi.")
        await state.update_data(products=[], index=0, sold=[], branch_id=branch_id)
        await message.answer("✅ Hisobotni yuboraymi? (ha/yo‘q)")
        await state.set_state(ReportState.confirm)
        return

    await state.update_data(products=products, index=0, sold=[], branch_id=branch_id)
    current = products[0]
    await message.answer(
        f"🛒 {current['product_name']} — nechta sotildi? ({current['unit']})\n"
        f"Raqam yuboring, masalan: 3 yoki 2.5"
    )
    await state.set_state(ReportState.product_loop)


@router.message(ReportState.product_loop)
async def process_products(message: types.Message, state: FSMContext):
    data = await state.get_data()
    products = data["products"]
    index = data["index"]
    sold = data["sold"]

    # Kiritilgan miqdor
    try:
        sold_amount = float(message.text.replace(" ", ""))
    except ValueError:
        await message.answer("❌ Faqat raqam kiriting, masalan: 5 yoki 2.5")
        return

    current = products[index]
    product_id = current["id"]
    old_qty = float(current["quantity"] or 0)
    new_qty = max(old_qty - sold_amount, 0)

    # Omborni avtomatik kamaytiramiz
    database.execute("UPDATE warehouse SET quantity=:q WHERE id=:id", {"q": new_qty, "id": product_id})

    # Sotuv ro‘yxatiga qo‘shamiz
    sold.append({
        "name": current["product_name"],
        "amount": sold_amount,
        "unit": current["unit"],
        "remaining": new_qty
    })

    # Keyingisi
    index += 1
    if index < len(products):
        next_p = products[index]
        await state.update_data(index=index, sold=sold)
        await message.answer(f"🛒 {next_p['product_name']} — nechta sotildi? ({next_p['unit']})")
    else:
        await state.update_data(sold=sold)
        await message.answer("✅ Hammasi yozildi! Hisobotni yuboraymi? (ha/yo‘q)")
        await state.set_state(ReportState.confirm)


@router.message(ReportState.confirm)
async def finish_report(message: types.Message, state: FSMContext):
    if message.text.lower() not in ("ha", "xa", "yes"):
        await message.answer("❌ Hisobot bekor qilindi.")
        await state.clear()
        return

    data = await state.get_data()
    user_id = message.from_user.id
    now = business_now()
    bdate = business_date(now)

    branch_id = data.get("branch_id")
    income = float(data.get("income", 0))
    expense = float(data.get("expense", 0))
    remaining_money = income - expense
    sold = data.get("sold", [])

    # Matnlar
    sold_text = "\n".join([f"- {s['name']} — {s['amount']} {s['unit']}" for s in sold]) or "—"
    remain_text = "\n".join([f"- {s['name']} — {s['remaining']} {s['unit']}" for s in sold]) or "—"

    # Bazaga yozamiz
    database.execute("""
        INSERT INTO reports (user_id, branch_id, date, income, expense, remaining, sold_items, notes)
        VALUES (:u, :b, :d, :i, :e, :r, :s, :n)
    """, {
        "u": user_id,
        "b": branch_id,
        "d": bdate,  # DATE tipida
        "i": income,
        "e": expense,
        "r": remaining_money,
        "s": sold_text,
        "n": remain_text
    })

    # Ishchiga yakuniy xabar
    pretty_day = bdate.strftime("%d.%m.%Y")
    await message.answer(
        f"🧾 <b>Bugungi hisobot ({pretty_day})</b>\n\n"
        f"💰 Daromad: <b>{fmt_sum(income)}</b> so‘m\n"
        f"💸 Rashod: <b>{fmt_sum(expense)}</b> so‘m\n"
        f"💵 Qolgan: <b>{fmt_sum(remaining_money)}</b> so‘m\n\n"
        f"🛒 <b>Sotilganlar:</b>\n{sold_text}\n\n"
        f"📦 <b>Omborda qoldi:</b>\n{remain_text}",
        parse_mode="HTML"
    )

    # Filial nomi
    branch_name = "-"
    if branch_id:
        b = database.fetchone("SELECT name FROM branches WHERE id=:id", {"id": branch_id})
        if b:
            branch_name = b["name"]

    # Superadmin(lar)ga yuborish (matn + Excel)
    admin_text = (
        f"📅 {pretty_day}\n"
        f"🏢 Filial: {branch_name}\n"
        f"👤 Ishchi: {message.from_user.full_name} (ID: {user_id})\n\n"
        f"💰 Daromad: {fmt_sum(income)} so‘m\n"
        f"💸 Rashod: {fmt_sum(expense)} so‘m\n"
        f"💵 Qolgan: {fmt_sum(remaining_money)} so‘m\n\n"
        f"🛒 Sotilganlar:\n{sold_text}\n\n"
        f"📦 Omborda qoldi:\n{remain_text}"
    )

    # Excel
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hisobot"

        headers = ["Nomi", "Miqdori", "Turi"]
        ws.append(headers)
        bold = Font(bold=True)
        for col in range(1, 4):
            ws.cell(row=1, column=col).font = bold
            ws.cell(row=1, column=col).fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Sotilganlar va qolganlar
        for s in sold:
            ws.append([s["name"], s["amount"], "Sotilgan"])
        for s in sold:
            ws.append([s["name"], s["remaining"], "Qolgan"])

        filename = f"hisobot_{branch_name}_{bdate}.xlsx".replace(" ", "_")
        path = os.path.join("/tmp", filename)
        wb.save(path)
    except Exception:
        path = None

    for admin_id in SUPERADMINS:
        try:
            await message.bot.send_message(admin_id, admin_text)
            if path and os.path.exists(path):
                await message.bot.send_document(admin_id, FSInputFile(path), caption="📊 Excel")
        except Exception:
            pass

    if path and os.path.exists(path):
        try:
            os.remove(path)
        except Exception:
            pass

    await state.clear()


# ===============================
# 📋 Ombor holati (faqat o‘z filialiga)
# ===============================
@router.message(F.text == "📋 Ombor holati")
async def show_warehouse(message: types.Message):
    user = database.fetchone("SELECT branch_id FROM users WHERE telegram_id=:tid", {"tid": message.from_user.id})
    if not user:
        await message.answer("⚠️ Sizning filial aniqlanmadi.")
        return

    branch_id = user["branch_id"]
    products = database.list_products_by_branch(branch_id)

    if not products:
        await message.answer("📦 Omborda hozircha mahsulotlar yo‘q.")
        return

    text = f"📋 <b>Ombor (filial: {branch_id})</b>\n\n"
    for p in products:
        qty = float(p["quantity"] or 0)
        qty_show = int(qty) if qty.is_integer() else qty
        text += f"• {p['product_name']} — {qty_show} {p['unit']}\n"

    await message.answer(text, parse_mode="HTML")


# =====================================
# 💰 BONUS / JARIMALAR — menyu va ko‘rinishlar
# =====================================
@router.message(F.text == "💰 Bonus / Jarimalarim")
async def open_bonus_menu(message: types.Message):
    ensure_bonus_tables()
    await message.answer("💰 Bonus yoki jarimalar bo‘limini tanlang:", reply_markup=get_bonus_kb())


@router.message(F.text == "📅 Bugungi")
async def show_today_bonus(message: types.Message):
    ensure_bonus_tables()
    today = business_date()
    user_id = message.from_user.id

    bonuses = database.fetchall("""
        SELECT amount, reason, created_at
        FROM bonuses
        WHERE user_id=:u AND DATE(created_at)=:d
        ORDER BY created_at DESC
    """, {"u": user_id, "d": today})

    fines = database.fetchall("""
        SELECT amount, reason, created_at
        FROM fines
        WHERE user_id=:u AND DATE(created_at)=:d
        ORDER BY created_at DESC
    """, {"u": user_id, "d": today})

    text = f"📅 <b>Bugungi ({today}) bonus va jarimalar:</b>\n\n"
    if not bonuses and not fines:
        text += "📭 Bugun yozuv yo‘q."
    else:
        if bonuses:
            text += "✅ <b>Bonuslar:</b>\n"
            for b in bonuses:
                text += f"➕ {fmt_sum(b['amount'])} — {b.get('reason') or '-'}\n"
            text += "\n"
        if fines:
            text += "❌ <b>Jarimalar:</b>\n"
            for f in fines:
                text += f"➖ {fmt_sum(f['amount'])} — {f.get('reason') or '-'}\n"

    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "📋 Umumiy")
async def show_all_bonus(message: types.Message):
    ensure_bonus_tables()
    user_id = message.from_user.id

    bonuses = database.fetchall("""
        SELECT amount, reason, created_at
        FROM bonuses
        WHERE user_id=:u
        ORDER BY created_at DESC
        LIMIT 30
    """, {"u": user_id})

    fines = database.fetchall("""
        SELECT amount, reason, created_at
        FROM fines
        WHERE user_id=:u
        ORDER BY created_at DESC
        LIMIT 30
    """, {"u": user_id})

    text = "📋 <b>So‘nggi 30 ta bonus/jarimalar:</b>\n\n"
    if not bonuses and not fines:
        text += "📭 Hozircha yozuv yo‘q."
    else:
        if bonuses:
            text += "✅ <b>Bonuslar:</b>\n"
            for b in bonuses:
                text += f"➕ {fmt_sum(b['amount'])} — {b.get('reason') or '-'}\n"
            text += "\n"
        if fines:
            text += "❌ <b>Jarimalar:</b>\n"
            for f in fines:
                text += f"➖ {fmt_sum(f['amount'])} — {f.get('reason') or '-'}\n"

    await message.answer(text, parse_mode="HTML")


# ===============================
# ⬅️ Orqaga — Asosiy menyuga
# ===============================
@router.message(F.text == "⬅️ Orqaga")
async def back_to_main_worker_menu(message: types.Message):
    await message.answer("🏠 Asosiy ishchi menyuga qaytdingiz:", reply_markup=get_worker_kb())
