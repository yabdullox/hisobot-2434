    # src/handlers/superadmin.py

from aiogram import Router, types, F
from aiogram.filters import Command
from aiogram.types import (
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    ReplyKeyboardMarkup,
    KeyboardButton,
    FSInputFile,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from keyboards.superadmin_kb import get_superadmin_kb
from config import SUPERADMIN_ID
from utils.excel_export import export_reports_to_excel
import database
import os

from datetime import datetime, date

router = Router()


# ===============================
# FSM holatlari (agar kerak bo'lsa keyin kengaytirish mumkin)
# ===============================
class AddBranchFSM(StatesGroup):
    name = State()
    branch_id = State()


class DelBranchFSM(StatesGroup):
    branch_id = State()


class AddAdminFSM(StatesGroup):
    name = State()
    phone = State()
    tg_id = State()
    branch_id = State()


class DelAdminFSM(StatesGroup):
    admin_id = State()


# ===============================
# Start (SuperAdmin tekshiruvi)
# ===============================
@router.message(Command("start"))
async def cmd_start(message: types.Message):
    # SUPERADMIN_ID — string yoki vergul bilan bo'lingan IDlar bo'lishi mumkin
    allowed = [s.strip() for s in str(SUPERADMIN_ID).split(",") if s.strip()]
    if str(message.from_user.id) not in allowed:
        await message.answer("❌ Siz SuperAdmin emassiz.")
        return

    await message.answer(
        "👋 Salom, SuperAdmin!\nHISOBOT24 boshqaruv paneli ishga tayyor.",
        reply_markup=get_superadmin_kb()
    )


# ===============================
# 👥 Adminni bir nechta filialga biriktirish (✅ belgili variant)
# ===============================
from aiogram.utils.keyboard import InlineKeyboardBuilder

# Holatlarni saqlash uchun FSM
from aiogram.fsm.state import State, StatesGroup

class AdminLinkState(StatesGroup):
    admin_id = State()
    selected_branches = State()


@router.message(F.text == "➕ Adminni filialga biriktirish")
async def start_admin_link(message: types.Message, state: FSMContext):
    """Admin tanlash uchun menyu chiqaradi."""
    admins = database.fetchall("SELECT id, full_name, telegram_id FROM users WHERE role='admin' ORDER BY id")
    if not admins:
        await message.answer("👥 Hozircha adminlar mavjud emas.")
        return

    kb = InlineKeyboardBuilder()
    for a in admins:
        kb.button(text=f"{a['full_name']} ({a['telegram_id']})", callback_data=f"choose_admin:{a['id']}")
    kb.button(text="❌ Bekor qilish", callback_data="cancel_link")
    await message.answer("👥 Qaysi adminni filial(larga) biriktirmoqchisiz?", reply_markup=kb.as_markup())


@router.callback_query(F.data.startswith("choose_admin:"))
async def choose_admin(callback: types.CallbackQuery, state: FSMContext):
    """Admin tanlangach — filiallar ro‘yxatini ko‘rsatadi."""
    admin_id = int(callback.data.split(":")[1])
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id")

    if not branches:
        await callback.message.answer("🏢 Hozircha filiallar mavjud emas.")
        await callback.answer()
        return

    await state.set_state(AdminLinkState.selected_branches)
    await state.update_data(admin_id=admin_id, selected_branches=[])

    await show_branch_selection(callback.message, state)
    await callback.answer()


async def show_branch_selection(message: types.Message, state: FSMContext):
    """Filiallar ro‘yxatini ✅ belgili qilib chiqaradi."""
    data = await state.get_data()
    selected = data.get("selected_branches", [])
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id")

    kb = InlineKeyboardBuilder()
    for b in branches:
        mark = "✅ " if b["id"] in selected else ""
        kb.button(text=f"{mark}{b['name']}", callback_data=f"toggle_branch:{b['id']}")

    kb.button(text="✅ Saqlash", callback_data="save_branches")
    kb.button(text="❌ Bekor qilish", callback_data="cancel_link")

    await message.edit_text("🏢 Filiallarni tanlang (5 tagacha). So‘ng ✅ Saqlash’ni bosing:", reply_markup=kb.as_markup())


@router.callback_query(F.data.startswith("toggle_branch:"))
async def toggle_branch(callback: types.CallbackQuery, state: FSMContext):
    """Filial tanlanganida ✅ belgini qo‘yish yoki olib tashlash."""
    branch_id = int(callback.data.split(":")[1])
    data = await state.get_data()
    selected = set(data.get("selected_branches", []))

    # Toggle holati
    if branch_id in selected:
        selected.remove(branch_id)
        msg = "❌ Tanlov olib tashlandi."
    else:
        if len(selected) >= 5:
            await callback.answer("⚠️ Faqat 5 ta filial tanlash mumkin!", show_alert=True)
            return
        selected.add(branch_id)
        msg = "✅ Tanlov qo‘shildi."

    await state.update_data(selected_branches=list(selected))
    await show_branch_selection(callback.message, state)
    await callback.answer(msg)


async def show_branch_selection(message: types.Message, state: FSMContext):
    """Filial ro‘yxatini chiqaradi — tanlanganlarda ✅ belgisi bor."""
    data = await state.get_data()
    selected = data.get("selected_branches", [])
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id")

    from aiogram.utils.keyboard import InlineKeyboardBuilder
    kb = InlineKeyboardBuilder()

    for b in branches:
        # Belgini faqat tanlangan filiallarga qo‘yamiz
        mark = "✅ " if b["id"] in selected else "🏢 "
        kb.button(text=f"{mark}{b['name']}", callback_data=f"toggle_branch:{b['id']}")

    kb.button(text="✅ Saqlash", callback_data="save_branches")
    kb.button(text="❌ Bekor qilish", callback_data="cancel_link")

    # Xabarni yangilaymiz (edit_text xatolik bermasligi uchun try-except)
    try:
        await message.edit_text(
            "🏢 Filiallarni tanlang (5 tagacha). So‘ng ✅ Saqlash’ni bosing:",
            reply_markup=kb.as_markup()
        )
    except Exception:
        await message.answer(
            "🏢 Filiallarni tanlang (5 tagacha). So‘ng ✅ Saqlash’ni bosing:",
            reply_markup=kb.as_markup()
        )

@router.callback_query(F.data == "cancel_link")
async def cancel_link(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("❌ Amal bekor qilindi.")
    await callback.answer()
# ===============================
# 📊 HISOBOTLAR BO‘LIMI
# ===============================
@router.message(F.text.in_(["📊 Bugungi hisobotlar", "📈 Umumiy hisobotlar"]))
async def choose_report_type(message: types.Message):
    """Hisobot turi tanlangandan keyin filiallar chiqadi."""
    action = "today" if message.text == "📊 Bugungi hisobotlar" else "all"

    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id ASC")
    if not branches:
        await message.answer("⚠️ Hozircha filiallar mavjud emas.")
        return

    # Inline tugmalar yaratamiz
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=b["name"], callback_data=f"{action}_branch:{b['id']}")]
            for b in branches
        ] + [
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel")]
        ]
    )

    await message.answer(
        f"📅 Qaysi filialning {'bugungi' if action == 'today' else 'umumiy'} hisobotlarini ko‘rmoqchisiz?",
        reply_markup=kb
    )


# ===============================
# ❌ BEKOR QILISH
# ===============================
@router.callback_query(F.data == "cancel")
async def cancel_action(callback: types.CallbackQuery):
    await callback.message.answer("❌ Bekor qilindi.", reply_markup=get_superadmin_kb())
    await callback.answer()





# ===============================
# 📋 FILIALLAR RO‘YXATI (INLINE)
# ===============================
@router.message(F.text.in_(["📊 Bugungi hisobotlar", "📈 Umumiy hisobotlar"]))
async def choose_report_type(message: types.Message):
    """Superadmin uchun filiallar ro‘yxatini ko‘rsatadi (bugungi hisobot uchun)."""
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id")
    if not branches:
        await message.answer("📭 Hozircha filiallar mavjud emas.")
        return

    keyboard = [
        [types.InlineKeyboardButton(text=f"🏢 {b['name']}", callback_data=f"today_branch:{b['id']}")]
        for b in branches
    ]
    await message.answer("📅 Qaysi filial uchun bugungi hisobotni ko‘rasiz?", 
                         reply_markup=types.InlineKeyboardMarkup(inline_keyboard=keyboard))


@router.message(F.text == "📈 Umumiy hisobotlar")
async def choose_branch_all(message: types.Message):
    """Superadmin uchun filiallar ro‘yxatini ko‘rsatadi (umumiy hisobot uchun)."""
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id")
    if not branches:
        await message.answer("📭 Hozircha filiallar mavjud emas.")
        return

    keyboard = [
        [types.InlineKeyboardButton(text=f"🏢 {b['name']}", callback_data=f"all_branch:{b['id']}")]
        for b in branches
    ]
    await message.answer("📈 Qaysi filial uchun umumiy hisobotni ko‘rasiz?", 
                         reply_markup=types.InlineKeyboardMarkup(inline_keyboard=keyboard))


# ===============================
# 📅 BUGUNGI HISOBOTLAR — TO‘G‘RILANGAN ISHLAYDIGAN VERSIYA
# ===============================
from datetime import datetime
import pytz

@router.callback_query(F.data.startswith("today_branch:"))
async def show_today_reports(callback: types.CallbackQuery):
    """Bugungi filial hisobotlarini ko‘rsatadi (O‘zbekiston vaqti bilan)."""
    branch_id = int(callback.data.split(":")[1])

    # 🕓 O‘zbekiston vaqti bilan hozirgi sana
    uz_tz = pytz.timezone("Asia/Tashkent")
    today = datetime.now(uz_tz).date()

    # 🧾 Hisobotlarni olish (sqlite uchun DATE() emas, to‘g‘ridan-to‘g‘ri qiymat bilan)
    reports = database.fetchall("""
        SELECT 
            r.user_id,
            u.full_name,
            u.branch_id,
            r.date,
            r.start_time,
            r.end_time,
            r.income,
            r.expense,
            r.remaining,
            r.sold_items,
            r.notes,
            b.name AS branch_name
        FROM reports r
        LEFT JOIN users u ON u.telegram_id = r.user_id
        LEFT JOIN branches b ON b.id = r.branch_id
        WHERE r.date = :today AND r.branch_id = :bid
        ORDER BY r.start_time
    """, {"today": str(today), "bid": branch_id})

    # ⚠️ Agar hisobot topilmasa
    if not reports:
        await callback.message.answer(
            f"📭 <b>Bugun ({today}) bu filialda hisobot topilmadi.</b>",
            parse_mode="HTML"
        )
        await callback.answer()
        return

    # 🏢 Filial nomini olish
    branch_name = reports[0]["branch_name"] or f"ID: {branch_id}"
    result = f"📅 <b>{branch_name}</b> — bugungi hisobotlar ({today}):\n\n"

    # 🧱 Har bir hisobotni formatlab chiqarish
    for r in reports:
        result += (
            f"👷‍♂️ <b>Ishchi:</b> {r['full_name'] or 'Noma’lum'}\n"
            f"🏢 <b>Filial:</b> {r['branch_name']} (ID: {r['branch_id']})\n"
            f"🆔 <b>Telegram ID:</b> <code>{r['user_id']}</code>\n\n"
            f"🕒 <b>Vaqt:</b> {r['start_time'] or '-'} — {r['end_time'] or '-'}\n"
            f"💰 <b>Daromad:</b> {fmt_sum(r['income'] or 0)} so‘m\n"
            f"💸 <b>Rashod:</b> {fmt_sum(r['expense'] or 0)} so‘m\n"
            f"💵 <b>Qolgan:</b> {fmt_sum(r['remaining'] or 0)} so‘m\n\n"
            f"🛒 <b>Sotilganlar:</b>\n{r['sold_items'] or '—'}\n\n"
            f"📦 <b>Qolgan mahsulotlar:</b>\n{r['notes'] or '—'}\n"
            "━━━━━━━━━━━━━━━━━━━━━━━\n"
        )

    # 📨 Natijani yuborish
    await callback.message.answer(result, parse_mode="HTML")
    await callback.answer()

# ===============================
# 📈 UMUMIY HISOBOTLAR
# ===============================
@router.callback_query(F.data.startswith("all_branch:"))
async def show_all_reports(callback: types.CallbackQuery):
    """Tanlangan filial uchun umumiy hisobotlar."""
    branch_id = int(callback.data.split(":")[1])

    reports = database.fetchall("""
        SELECT 
            r.user_id,
            u.full_name,
            u.branch_id,
            r.date,
            r.start_time,
            r.end_time,
            r.income,
            r.expense,
            r.remaining,
            r.sold_items,
            r.notes,
            b.name AS branch_name
        FROM reports r
        LEFT JOIN users u ON u.telegram_id = r.user_id
        LEFT JOIN branches b ON b.id = r.branch_id
        WHERE r.branch_id = :bid
        ORDER BY r.date DESC, r.start_time
        LIMIT 50
    """, {"bid": branch_id})

    if not reports:
        await callback.message.answer("📭 Bu filialda hali hisobotlar yo‘q.")
        await callback.answer()
        return

    branch_name = reports[0]["branch_name"] or f"ID: {branch_id}"
    result = f"📈 <b>{branch_name}</b> — umumiy hisobotlar:\n\n"

    for r in reports:
        result += (
            f"📅 <b>Sana:</b> {r['date']}\n"
            f"👷‍♂️ <b>Ishchi:</b> {r['full_name'] or 'Noma’lum'}\n"
            f"🕒 {r['start_time'] or '-'} — {r['end_time'] or '-'}\n"
            f"💰 {fmt_sum(r['income'])} | 💸 {fmt_sum(r['expense'])} | 💵 {fmt_sum(r['remaining'])}\n"
            f"🛒 Sotilganlar:\n{r['sold_items'] or '—'}\n\n"
            f"📦 Omborda:\n{r['notes'] or '—'}\n"
            "────────────────────────────\n"
        )

    await callback.message.answer(result, parse_mode="HTML")
    await callback.answer()


# ===============================
# 🔢 SUM FORMAT FUNKSIYA
# ===============================
def fmt_sum(v):
    try:
        return f"{float(v):,.0f}".replace(",", " ")
    except:
        return str(v)
# ===============================
# Filiallar ro'yxati, qo'shish, o'chirish
# ===============================
@router.message(F.text == "🏢 Filiallar ro‘yxati")
async def branch_list(message: types.Message):
    rows = database.fetchall("SELECT id, name FROM branches ORDER BY id ASC")
    if not rows:
        await message.answer("🏢 Hozircha filiallar mavjud emas.")
        return
    text = "🏢 Filiallar ro‘yxati:\n\n"
    for r in rows:
        text += f"🆔 {r['id']} — {r['name']}\n"
    await message.answer(text)


@router.message(F.text == "➕ Filial qo‘shish")
async def add_branch_start(message: types.Message, state: FSMContext):
    await state.set_state(AddBranchFSM.name)
    await message.answer("🏢 Filial nomini kiriting:")


@router.message(AddBranchFSM.name)
async def add_branch_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(AddBranchFSM.branch_id)
    await message.answer("🆔 Filial ID raqamini kiriting (faqat raqam):")


@router.message(AddBranchFSM.branch_id)
async def add_branch_finish(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("❗️Faqat raqam kiriting.")
        return
    data = await state.get_data()
    database.execute("INSERT INTO branches (id, name) VALUES (:id, :name)", {"id": int(message.text), "name": data["name"]})
    await state.clear()
    await message.answer(f"✅ Filial qo‘shildi: {data['name']} (ID: {message.text})")


@router.message(F.text == "❌ Filialni o‘chirish")
async def del_branch_start(message: types.Message, state: FSMContext):
    rows = database.fetchall("SELECT id, name FROM branches")
    if not rows:
        await message.answer("❌ Filiallar mavjud emas.")
        return
    text = "🗑️ Qaysi filialni o‘chirmoqchisiz?\n\n"
    for r in rows:
        text += f"{r['id']}. {r['name']}\n"
    await message.answer(text)
    await state.set_state(DelBranchFSM.branch_id)


@router.message(DelBranchFSM.branch_id)
async def del_branch_finish(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("❗️Filial ID raqamini kiriting.")
        return
    database.execute("DELETE FROM branches WHERE id=:id", {"id": int(message.text)})
    await state.clear()
    await message.answer("✅ Filial muvaffaqiyatli o‘chirildi.")


# ===============================
# Adminlar ro'yxati va qo'shish/o'chirish
# ===============================
# @router.message(F.text == "👥 Adminlar ro‘yxati")
# async def admin_list(message: types.Message):
#     admins = database.fetchall("SELECT id, full_name, telegram_id, branch_id FROM users WHERE role='admin'")
#     if not admins:
#         await message.answer("👥 Adminlar hozircha mavjud emas.")
#         return
#     text = "👥 Adminlar:\n\n"
#     for a in admins:
#         text += f"{a['id']}. {a['full_name']} — 🆔 {a['telegram_id']} — Filial: {a.get('branch_id','—')}\n"
#     await message.answer(text)

# ===============================
# 👥 ADMINLAR RO‘YXATI
# ===============================
@router.message(F.text == "👥 Adminlar ro‘yxati")
async def show_admin_list(message: types.Message):
    """Barcha adminlarni chiroyli ro‘yxatda ko‘rsatadi"""
    admins = database.fetchall("""
        SELECT 
            u.id,
            u.full_name,
            u.telegram_id,
            COALESCE(b.name, CONCAT('Filial ID: ', u.branch_id)) AS branch_name,
            TO_CHAR(u.created_at, 'YYYY-MM-DD HH24:MI') AS created_at
        FROM users u
        LEFT JOIN branches b ON b.id = u.branch_id
        WHERE LOWER(COALESCE(u.role, '')) LIKE '%admin%'
        ORDER BY u.id ASC
    """)

    if not admins:
        await message.answer("👥 Adminlar hozircha mavjud emas.")
        return

    header = "👥 <b>Adminlar ro‘yxati:</b>\n\n"
    text = header
    count = 0
    all_texts = []

    for idx, a in enumerate(admins, start=1):
        block = (
            f"<b>{idx}.</b> 👤 {a['full_name'] or '-'}\n"
            f"🆔 <b>ID:</b> <code>{a['id']}</code>\n"
            f"💬 <b>Telegram ID:</b> <code>{a['telegram_id'] or '-'}</code>\n"
            f"🏢 <b>Filial:</b> {a['branch_name']}\n"
            f"🕓 <b>Qo‘shilgan:</b> {a['created_at'] or '-'}\n"
            "━━━━━━━━━━━━━━━━━━━━━━━\n"
        )

        if len(text) + len(block) > 3500:
            all_texts.append(text)
            text = ""

        text += block
        count += 1

    if text:
        all_texts.append(text)

    for part in all_texts:
        await message.answer(part, parse_mode="HTML")

    await message.answer(f"✅ Jami <b>{count}</b> ta admin topildi.", parse_mode="HTML")


# ===============================
# ➕ ADMIN QO‘SHISH
# ===============================
@router.message(F.text == "➕ Admin qo‘shish")
async def add_admin_start(message: types.Message, state: FSMContext):
    await state.set_state(AddAdminFSM.name)
    await message.answer("👤 Admin to‘liq ismini kiriting:")


@router.message(AddAdminFSM.name)
async def add_admin_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(AddAdminFSM.phone)
    await message.answer("📞 Admin telefon raqamini kiriting:")


@router.message(AddAdminFSM.phone)
async def add_admin_phone(message: types.Message, state: FSMContext):
    await state.update_data(phone=message.text)
    await state.set_state(AddAdminFSM.tg_id)
    await message.answer("🆔 Admin Telegram ID’sini kiriting:")


@router.message(AddAdminFSM.tg_id)
async def add_admin_tgid(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("❗️ Telegram ID raqam bo‘lishi kerak.")
        return
    await state.update_data(tg_id=int(message.text))
    await state.set_state(AddAdminFSM.branch_id)
    await message.answer("🏢 Admin qaysi filialga biriktiriladi? Filial ID kiriting:")


@router.message(AddAdminFSM.branch_id)
async def add_admin_finish(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("❗️ Filial ID faqat raqam bo‘lishi kerak.")
        return

    data = await state.get_data()
    full_name = f"{data['name']} ({data['phone']})"

    # Admin avvaldan mavjudmi — tekshirish
    exists = database.fetchone("SELECT id FROM users WHERE telegram_id = :tg_id", {"tg_id": data["tg_id"]})
    if exists:
        await message.answer("⚠️ Bu Telegram ID allaqachon tizimda mavjud.")
        await state.clear()
        return

    database.execute("""
        INSERT INTO users (telegram_id, full_name, role, branch_id, created_at)
        VALUES (:tg, :fn, 'admin', :br, :dt)
    """, {
        "tg": data["tg_id"],
        "fn": full_name,
        "br": int(message.text),
        "dt": datetime.now()
    })

    await state.clear()
    await message.answer(f"✅ Admin qo‘shildi:\n👤 {full_name}\n🏢 Filial ID: {message.text}")


# ===============================
# 🗑️ ADMINNI O‘CHIRISH
# ===============================
@router.message(F.text == "🗑️ Adminni o‘chirish")
async def del_admin_start(message: types.Message, state: FSMContext):
    admins = database.fetchall("""
        SELECT 
            u.id,
            u.full_name,
            u.telegram_id,
            COALESCE(b.name, CONCAT('Filial ID: ', u.branch_id)) AS branch_name,
            TO_CHAR(u.created_at, 'YYYY-MM-DD HH24:MI') AS created_at
        FROM users u
        LEFT JOIN branches b ON b.id = u.branch_id
        WHERE LOWER(COALESCE(u.role, '')) LIKE '%admin%'
        ORDER BY u.id ASC
    """)

    if not admins:
        await message.answer("👥 Adminlar hozircha mavjud emas.")
        return

    text_header = "🗑️ <b>Adminlar ro‘yxati (o‘chirish uchun):</b>\n\n"
    text = text_header
    all_texts = []

    for idx, a in enumerate(admins, start=1):
        block = (
            f"<b>{idx}.</b> 👤 {a['full_name'] or '-'}\n"
            f"🆔 <b>ID:</b> <code>{a['id']}</code>\n"
            f"💬 <b>Telegram ID:</b> <code>{a['telegram_id'] or '-'}</code>\n"
            f"🏢 <b>Filial:</b> {a['branch_name']}\n"
            f"🕓 <b>Qo‘shilgan:</b> {a['created_at'] or '-'}\n"
            "━━━━━━━━━━━━━━━━━━━━━━━\n"
        )

        if len(text) + len(block) > 3500:
            all_texts.append(text)
            text = ""

        text += block

    if text:
        all_texts.append(text)

    for part in all_texts:
        await message.answer(part, parse_mode="HTML")

    await message.answer(
        "✏️ O‘chirish uchun admin ID raqamini kiriting:",
        parse_mode="HTML"
    )
    await state.set_state(DelAdminFSM.admin_id)


@router.message(DelAdminFSM.admin_id)
async def del_admin_finish(message: types.Message, state: FSMContext):
    admin_id = message.text.strip()
    if not admin_id.isdigit():
        await message.answer("❗️ Faqat raqam kiriting (admin ID).")
        return

    admin = database.fetchone("SELECT * FROM users WHERE id = :id AND LOWER(role) LIKE '%admin%'", {"id": int(admin_id)})
    if not admin:
        await message.answer("⚠️ Bunday ID raqamli admin topilmadi.")
        return

    database.execute("DELETE FROM users WHERE id = :id", {"id": int(admin_id)})
    await state.clear()
    await message.answer(f"✅ Admin (ID: <b>{admin_id}</b>) muvaffaqiyatli o‘chirildi.", parse_mode="HTML")

# ===============================
# Export menyulari (Excel/CSV)
# ===============================
# =====================================
# 📤 Export menyusi
# =====================================
@router.message(F.text == "📤 Export (Excel / CSV)")
async def export_menu(message: types.Message):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Bugungi hisobotni eksport qilish", callback_data="export:today"),
        ],
        [
            InlineKeyboardButton(text="📆 Umumiy barcha hisobotlar (Excel)", callback_data="export:all")
        ],
        [
            InlineKeyboardButton(text="❌ Bekor qilish", callback_data="export:cancel")
        ]
    ])
    await message.answer("📊 Qaysi turdagi hisobotni eksport qilmoqchisiz?", reply_markup=kb)


# =====================================
# 📆 Umumiy barcha hisobotlarni Excel’ga eksport qilish
# =====================================
@router.callback_query(F.data == "export:all")
async def export_all_reports(callback: types.CallbackQuery):
    try:
        reports = database.fetchall("""
            SELECT 
                u.full_name AS worker_name,
                u.branch_id,
                b.name AS branch_name,
                r.date,
                r.start_time,
                r.end_time,
                r.text AS report_text
            FROM reports r
            JOIN users u ON r.user_id = u.telegram_id
            JOIN branches b ON u.branch_id = b.id
            ORDER BY b.id, r.date DESC
        """)

        if not reports:
            await callback.message.answer("📭 Umumiy hisobotlar mavjud emas.")
            return

        wb = Workbook()
        sheet_map = {}

        # Excel dizayn
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        for r in reports:
            branch_name = r["branch_name"] or f"Filial_{r['branch_id']}"
            sheet_name = branch_name[:31]
            if sheet_name not in sheet_map:
                ws = wb.create_sheet(title=sheet_name)
                sheet_map[sheet_name] = ws
                headers = ["👷 Ishchi", "📅 Sana", "🕘 Boshlanish", "🏁 Tugash", "🧾 Hisobot matni"]
                ws.append(headers)
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=c)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = border

            ws = sheet_map[sheet_name]
            ws.append([
                r["worker_name"],
                r["date"].strftime("%Y-%m-%d") if r["date"] else "-",
                r["start_time"] or "-",
                r["end_time"] or "-",
                r["report_text"] or "",
            ])

        # Sheet kengligi
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 4

        # Default bo'sh "Sheet"ni o'chiramiz
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # ⚙️ Fayl nomini yaratamiz (datetime moduli bilan to‘g‘ri ishlaydi)
        from datetime import datetime
        filename = f"Umumiy_Hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        path = os.path.join("/tmp", filename)
        wb.save(path)

        await callback.message.answer_document(
            FSInputFile(path),
            caption="📊 <b>Barcha filiallar bo‘yicha umumiy Excel fayl tayyor!</b>",
            parse_mode="HTML"
        )
        os.remove(path)

    except Exception as e:
        await callback.message.answer(f"⚠️ Xatolik: {e}")


# =====================================
# 📅 Bugungi hisobotlar uchun filial tanlash
# =====================================
@router.callback_query(F.data == "export:today")
async def choose_branch_today(callback: types.CallbackQuery):
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id ASC")
    if not branches:
        await callback.message.answer("🏢 Hech qanday filial topilmadi.")
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=b["name"], callback_data=f"export_branch:{b['id']}")] for b in branches
    ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="export:cancel")])

    await callback.message.answer("📍 Qaysi filialning bugungi hisobotini eksport qilmoqchisiz?", reply_markup=kb)


# =====================================
# 📍 Tanlangan filial uchun bugungi hisobotni Excel’ga eksport qilish
# =====================================
@router.callback_query(F.data.startswith("export_branch:"))
async def export_today_branch(callback: types.CallbackQuery):
    branch_id = int(callback.data.split(":")[1])
    from datetime import date
    today = date.today()

    reports = database.fetchall("""
        SELECT 
            u.full_name AS worker_name,
            r.date,
            r.start_time,
            r.end_time,
            r.text AS report_text,
            b.name AS branch_name
        FROM reports r
        JOIN users u ON r.user_id = u.telegram_id
        JOIN branches b ON u.branch_id = b.id
        WHERE r.date = :today AND u.branch_id = :bid
        ORDER BY r.start_time ASC
    """, {"today": today, "bid": branch_id})

    if not reports:
        await callback.message.answer("📭 Bugun bu filialda hisobot yo‘q.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = reports[0]["branch_name"]

    headers = ["👷 Ishchi", "📅 Sana", "🕘 Boshlanish", "🏁 Tugash", "🧾 Hisobot matni"]
    ws.append(headers)

    # Dizayn
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    for r in reports:
        ws.append([
            r["worker_name"],
            r["date"].strftime("%Y-%m-%d"),
            r["start_time"] or "-",
            r["end_time"] or "-",
            r["report_text"] or "",
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    from datetime import datetime
    filename = f"{reports[0]['branch_name']}_Bugungi_{datetime.now().strftime('%Y%m%d')}.xlsx"
    path = os.path.join("/tmp", filename)
    wb.save(path)

    await callback.message.answer_document(
        FSInputFile(path),
        caption=f"📅 <b>{reports[0]['branch_name']}</b> filialining bugungi hisobotlari tayyor ✅",
        parse_mode="HTML"
    )
    os.remove(path)
# =====================================
# 📈 Umumiy barcha hisobotlar — bitta filial uchun
# =====================================
@router.callback_query(F.data == "export:all")
async def choose_branch_all(callback: types.CallbackQuery):
    """Filial tanlash — umumiy hisobot uchun."""
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id ASC")
    if not branches:
        await callback.message.answer("🏢 Hech qanday filial topilmadi.")
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=b["name"], callback_data=f"export_all_branch:{b['id']}")] for b in branches
    ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="export:cancel")])

    await callback.message.answer("📊 Qaysi filialning <b>umumiy hisobotini</b> eksport qilmoqchisiz?", 
                                  reply_markup=kb, parse_mode="HTML")


# =====================================
# 📆 Umumiy barcha hisobotlar — filial tanlash (xuddi bugungidek)
# =====================================
@router.callback_query(F.data == "export:all")
async def choose_branch_all(callback: types.CallbackQuery):
    """Umumiy hisobotlar uchun filial tanlash."""
    branches = database.fetchall("SELECT id, name FROM branches ORDER BY id ASC")
    if not branches:
        await callback.message.answer("🏢 Hech qanday filial topilmadi.")
        return

    # Inline tugmalar (bir xil dizayn)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=b["name"], callback_data=f"export_all_branch:{b['id']}")] for b in branches
    ])
    kb.inline_keyboard.append([InlineKeyboardButton(text="❌ Bekor qilish", callback_data="export:cancel")])

    await callback.message.answer(
        "📊 Qaysi filialning <b>umumiy hisobotini</b> eksport qilmoqchisiz?",
        reply_markup=kb,
        parse_mode="HTML"
    )


# =====================================
# 📈 Tanlangan filial uchun umumiy hisobotni Excel’ga eksport qilish
# =====================================
@router.callback_query(F.data.startswith("export_all_branch:"))
async def export_all_branch_reports(callback: types.CallbackQuery):
    """Tanlangan filialning barcha hisobotlarini Excel'ga chiqaradi."""
    branch_id = int(callback.data.split(":")[1])

    reports = database.fetchall("""
        SELECT 
            u.full_name AS worker_name,
            r.date,
            r.start_time,
            r.end_time,
            r.text AS report_text,
            b.name AS branch_name
        FROM reports r
        JOIN users u ON r.user_id = u.telegram_id
        JOIN branches b ON u.branch_id = b.id
        WHERE u.branch_id = :bid AND r.text IS NOT NULL
        ORDER BY r.date DESC, r.start_time
    """, {"bid": branch_id})

    if not reports:
        await callback.message.answer("📭 Bu filialda hali umumiy hisobotlar yo‘q.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    import os
    from aiogram.types import FSInputFile

    wb = Workbook()
    ws = wb.active
    ws.title = reports[0]["branch_name"]

    # Headerlar
    headers = ["👷 Ishchi", "📅 Sana", "🕘 Boshlanish", "🏁 Tugash", "🧾 Hisobot matni"]
    ws.append(headers)

    # Header dizayni
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border

    # Ma'lumotlarni qo‘shish
    for r in reports:
        ws.append([
            r["worker_name"],
            r["date"].strftime("%Y-%m-%d"),
            r["start_time"] or "-",
            r["end_time"] or "-",
            r["report_text"] or "",
        ])

    # Ustun kengligi avtomatik
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    filename = f"{reports[0]['branch_name']}_Umumiy_{datetime.now().strftime('%Y%m%d')}.xlsx"
    path = os.path.join("/tmp", filename)
    wb.save(path)

    await callback.message.answer_document(
        FSInputFile(path),
        caption=f"📈 <b>{reports[0]['branch_name']}</b> filialining umumiy hisobotlari tayyor ✅",
        parse_mode="HTML"
    )
    os.remove(path)

# =====================================
# ❌ Bekor qilish
# =====================================
@router.callback_query(F.data == "export:cancel")
async def cancel_export(callback: types.CallbackQuery):
    await callback.message.answer("❌ Bekor qilindi.")
# ===============================
# Bonus/Jarimalar ro'yxati
# ===============================
@router.message(F.text == "💰 Bonus/Jarimalar ro‘yxati")
async def show_all_fines_and_bonuses(message: types.Message):
    fines = database.fetchall(
        """
        SELECT u.full_name, f.amount, f.reason, f.created_at
        FROM fines f
        LEFT JOIN users u ON f.user_id = u.telegram_id
        ORDER BY f.created_at DESC
        LIMIT 20
        """
    )
    bonuses = database.fetchall(
        """
        SELECT u.full_name, b.amount, b.reason, b.created_at
        FROM bonuses b
        LEFT JOIN users u ON b.user_id = u.telegram_id
        ORDER BY b.created_at DESC
        LIMIT 20
        """
    )

    text = "💰 <b>So‘nggi 20 ta Bonus va Jarimalar</b>\n\n"
    if not fines and not bonuses:
        await message.answer("📂 Bonus yoki jarima yozuvlari topilmadi.")
        return

    if bonuses:
        text += "✅ <b>Bonuslar:</b>\n"
        for b in bonuses:
            text += f"👤 {b['full_name']} | +{b['amount']:,} so‘m\n📅 {b['created_at']}\n📝 {b['reason']}\n\n"
    if fines:
        text += "❌ <b>Jarimalar:</b>\n"
        for f in fines:
            text += f"👤 {f['full_name']} | -{f['amount']:,} so‘m\n📅 {f['created_at']}\n📝 {f['reason']}\n\n"

    await message.answer(text, parse_mode="HTML")


# ===============================
# Menyuga qaytish
# ===============================
@router.message(F.text == "⬅️ Menyuga qaytish")
async def back_to_menu(message: types.Message):
    allowed = [s.strip() for s in str(SUPERADMIN_ID).split(",") if s.strip()]
    if str(message.from_user.id) not in allowed:
        await message.answer("❌ Siz SuperAdmin emassiz.")
        return
    await message.answer("🏠 Asosiy menyuga qaytdingiz.", reply_markup=get_superadmin_kb())
